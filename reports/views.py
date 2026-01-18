from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from inventory.models import Laptop, Assessment


@login_required
def report_list(request):
    """Halaman preview laporan"""
    laptops = Laptop.objects.select_related('assessment').all()
    
    # Statistics
    total_assets = laptops.count()
    grade_a = Assessment.objects.filter(final_grade='A').count()
    grade_b = Assessment.objects.filter(final_grade='B').count()
    grade_cd = Assessment.objects.filter(final_grade__in=['C', 'D']).count()
    
    context = {
        'laptops': laptops,
        'total_assets': total_assets,
        'grade_a': grade_a,
        'grade_b': grade_b,
        'grade_cd': grade_cd,
    }
    return render(request, 'reports/report_list.html', context)


@login_required
def generate_pdf(request):
    """Generate PDF report"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO
    from datetime import datetime
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    # Header
    elements.append(Paragraph("PROTELINDO", title_style))
    elements.append(Paragraph("PT Profesional Telekomunikasi Indonesia", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Report title
    elements.append(Paragraph("LAPORAN REKAPITULASI KONDISI ASET", title_style))
    elements.append(Paragraph(f"Tanggal: {datetime.now().strftime('%d %B %Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Get data
    laptops = Laptop.objects.select_related('assessment').all()
    
    # Statistics table
    stats_data = [
        ['Total Aset', 'Grade A', 'Grade B', 'Grade C & D'],
        [
            str(laptops.count()),
            str(Assessment.objects.filter(final_grade='A').count()),
            str(Assessment.objects.filter(final_grade='B').count()),
            str(Assessment.objects.filter(final_grade__in=['C', 'D']).count()),
        ]
    ]
    
    stats_table = Table(stats_data, colWidths=[3*cm, 3*cm, 3*cm, 3*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))
    
    # Main data table
    table_data = [['No', 'Kode Aset', 'Model Aset', 'User Terakhir', 'Status', 'Kondisi', 'Grade']]
    
    for idx, laptop in enumerate(laptops, 1):
        assessment = getattr(laptop, 'assessment', None)
        condition = ', '.join(assessment.get_condition_tags()) if assessment else '-'
        grade = assessment.final_grade if assessment else '-'
        
        table_data.append([
            str(idx),
            laptop.asset_code,
            f"{laptop.brand} {laptop.model}",
            laptop.last_user,
            laptop.get_status_display(),
            condition[:30],
            grade
        ])
    
    main_table = Table(table_data, colWidths=[1*cm, 2*cm, 4*cm, 3*cm, 2.5*cm, 3.5*cm, 1.5*cm])
    main_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(main_table)
    elements.append(Spacer(1, 40))
    
    # Signature area
    sig_data = [
        ['Admin Gudang', '', 'Kepala Aset IT'],
        ['', '', ''],
        ['', '', ''],
        ['Nama & Tanggal', '', 'Nama & Tanggal'],
    ]
    sig_table = Table(sig_data, colWidths=[6*cm, 4*cm, 6*cm])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LINEBELOW', (0, 3), (0, 3), 1, colors.black),
        ('LINEBELOW', (2, 3), (2, 3), 1, colors.black),
    ]))
    elements.append(sig_table)
    
    # Build PDF
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Laporan_Aset_{datetime.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    return response


@login_required
def export_excel(request):
    """Export data to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from datetime import datetime
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Aset"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "LAPORAN REKAPITULASI KONDISI ASET"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Tanggal: {datetime.now().strftime('%d %B %Y')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['No', 'Kode Aset', 'Model Aset', 'User Terakhir', 'Status', 'Kondisi', 'Grade']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    laptops = Laptop.objects.select_related('assessment').all()
    for row_idx, laptop in enumerate(laptops, 5):
        assessment = getattr(laptop, 'assessment', None)
        condition = ', '.join(assessment.get_condition_tags()) if assessment else '-'
        grade = assessment.final_grade if assessment else '-'
        
        data = [
            row_idx - 4,
            laptop.asset_code,
            f"{laptop.brand} {laptop.model}",
            laptop.last_user,
            laptop.get_status_display(),
            condition,
            grade
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col in [1, 7] else "left")
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 10
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Laporan_Aset_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    return response
